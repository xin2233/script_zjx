#!/usr/bin/env python
# -*- coding: utf-8 -*-

#导入模块 openpyxl
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

#读取excel文件
#括号中的字符串为你要比较的两个excel的路径，注意用“/”
wb_a = openpyxl.load_workbook('d:/test.xlsx')
wb_b = openpyxl.load_workbook('d:/test2.xlsx')
#定义一个方法来获取表格中某一列的内容，返回一个列表
#在这里，我的表格中：IP是具有唯一性的，所以我用它来区分数据的差异，而IP这一列在我的表格中是第“G”列
def getIP(wb):

    sheet = wb.get_sheet_by_name('Sheet1')
    # sheet = wb.get_active_sheet()
    ip = []
    for cellobj in sheet['A']:
        ip.append(cellobj.value)

    return ip
#获得ip列表
ip_a = getIP(wb_a)
ip_b = getIP(wb_b)
#将两个列表转换成集合
aa = set(ip_a)
bb = set(ip_b)
#找出两个列表的不同行，并转换成列表
difference = list(aa ^ bb)
#打印出列表中的元素
#到这一步，两个表格中不同的数据已经被找出来了
for i in difference:
    print (i)

#将不同行高亮显示
print ("开始第一张表" + "----" *10)
a = wb_a.get_sheet_by_name('Sheet1')
for cellobj in a:
    if cellobj.value in difference:
        print (cellobj.value)
        cellobj.font = Font(color=colors.BLACK, italic=True ,bold = True)
        cellobj.fill = PatternFill("solid", fgColor="DDDDDD")
print ("开始第二张表" + "----" *10)
b = wb_b.get_sheet_by_name('Sheet1')
for cellobj in b:
    if cellobj.value in difference:
        print (cellobj.value)
        cellobj.font = Font(color=colors.BLACK, italic=True ,bold = True)
        cellobj.fill = PatternFill("solid", fgColor="DDDDDD")

wb_a.save('d:/a.xlsx')
wb_b.save('d:/b.xlsx')